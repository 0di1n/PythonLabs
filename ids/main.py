import openpyxl

students = {
    "Vitaly_Prikhodko": [12, 10, 9, 10, 5, 7, 8, 5, 12, 10], 
    "Dmytro_Kropyvnytskyi": [12, 10, 9, 5, 6, 7, 4, 3, 12, 4],
    "Mikhail_Romanenko": [12, 3, 4, 6, 5, 5, 3, 7, 5, 4], 
    "Maxim_Derizemlya": [12, 4, 10, 7, 5, 8, 3, 3, 5, 7], 
    "Victoria_Zhuk": [10, 10, 10, 9, 10, 9, 10, 8, 7 , 12], 
    "Andrey_Kuryanov": [5, 6, 7, 5, 4, 7, 5, 4, 4, 8], 
    "Oksana_Dubovets": [7, 8, 5, 8, 8, 9, 8, 7, 7, 10], 
    "Nikita_Stroganov": [6, 7, 8, 9, 10, 10 , 10, 10, 10 ,9], 
    "Karina_Nikolaenko": [2, 3, 5, 4, 5, 4, 3, 3, 5 ,8], 
    "Eugenia_Dron": [12, 12, 12, 10, 10, 9, 8, 9, 9 , 8]}


# Функція для додавання студента до словнику
def addStudent(Name: str, Grades: []):
    try:
        if name in students:
            raise ValueError("Учень з таким ім'ям вже існує.")
        students[name] = grades
    except ValueError as e:
        print(e)

# Функція для видалення студента зі словника
def removeStudent(Name: str):
    try:
        if name not in students:
            raise ValueError("Учень з таким ім'ям не існує.")
        del students[name]
    except ValueError as e:
        print(e)

# Функція для знаходження середньої оцінки
def calculateAverage():
    total_grades = sum(sum(grades) for grades in students.values())
    total_students = sum(len(grades) for grades in students.values())
    if total_students == 0:
        return 0.0
    classAverage = total_grades / total_students
    return classAverage

# Функція для виводу середніх оцінок учнів
def displayStudentsAverage():
    for name, grades in students.items():
        average = sum(grades) / len(grades)
        print(f"{name}: {average:.2f}")

# Функція для знаходження студентів з середньою оцінкою вище чим середня класу
def aboveClassAverage():
    classAverage = calculateAverage()
    aboveAverageStudents = [name for name, grades in students.items() if sum(grades) / len(grades) > classAverage]
    return aboveAverageStudents

# Функція для виводу імен та оцінок
def displayStudents():
    for name, grades in students.items():
        print(f"{name}: {grades}")

# Функція для виводу відсортованих імен та оцінок
def displaySortedKeys():
    sortedKeys = sorted(students.keys())
    for name in sortedKeys:
        print(f"{name}: {grades}")


# Створення воркбуку
workbook = openpyxl.Workbook()

# Вибір активного листа
sheet = workbook.active

while True:
    print("\nОберіть опцію:")
    print("1. Додати учня")
    print("2. Видалити учня")
    print("3. Порахувати середню оцінку")
    print("4. Показати середню оцінку учнів")
    print("5. Показати учнів з вищою середньою")
    print("6. Показати всіх учнів")
    print("7. Показати відсортовані імена учнів")
    print("8. Експорт в excel таблицю")
    print("q. Вийти")
    choice = input("Ваш вибір: ")
    if choice == "1":
        name = input("Введіть ім'я учня: ")
        grades = [int(grade) for grade in input("Введіть оцінки через пробіл: ").split()]
        addStudent(name, grades)
    elif choice == "2":
        name = input("Введіть ім'я учня, якого потрібно видалити: ")
        removeStudent(name)
    elif choice == "3":
        classAverage = calculateAverage()
        print(f"Середня оцінка по класу: {classAverage:.2f}")
    elif choice == "4":
        displayStudentsAverage()
    elif choice == "5":
        aboveAverageStudents = aboveClassAverage()
        print("Учні з вищою середньою оцінкою в класі:")
        for student in aboveAverageStudents:
            print(student)
    elif choice == "6":
        print("Учні у класі:")
        displayStudents()
    elif choice == "7":
        print("Всі імена учнів відсортовані:")
        displaySortedKeys()
    elif choice == "8":
        name = input("Зберегти в файл: ")
        sheet['A1'] = 'Студент'
        sheet['B1'] = 'Оцінки'
        for i in range(1, 11):
            sheet.cell(row=1, column=i+1, value=f'Оцінка {i}')
        row_number = 2
        for student, grades in students.items():
            # Записати імя студента
            sheet.cell(row=row_number, column=1, value=student)

            # Записати оцінки в кліткинки
            for col_num, grade in enumerate(grades, start=2):
                sheet.cell(row=row_number, column=col_num, value=grade)

            row_number += 1
        try:
            workbook.save(name)
        except:
            print("Не можу зберегти файл!")
    elif choice == "q":
        break